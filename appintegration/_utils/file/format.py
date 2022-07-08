from typing import List, Tuple, Iterable, Union
from abc import ABCMeta, ABC, abstractmethod

try:
    # It should install this package if user want to run the object *XLSXFormat*
    # command line: pip install openpyxl
    from openpyxl import load_workbook, Workbook
except ImportError:
    pass

from xml.etree.ElementTree import Element
import xml.etree.ElementTree as ET
import configparser
import json
import csv



class BaseFile(metaclass=ABCMeta):

    @property
    @abstractmethod
    def file_path(self) -> str:
        pass


    @file_path.setter
    @abstractmethod
    def file_path(self, path: str) -> None:
        pass


    @property
    @abstractmethod
    def mode(self) -> str:
        pass


    @mode.setter
    @abstractmethod
    def mode(self, mode: str) -> None:
        pass


    @property
    @abstractmethod
    def encoding(self) -> str:
        pass


    @encoding.setter
    @abstractmethod
    def encoding(self, encoding: str) -> None:
        pass


    @abstractmethod
    def open(self) -> None:
        pass


    @abstractmethod
    def write(self, data: Iterable[Iterable]) -> None:
        pass


    @abstractmethod
    def read(self, *args, **kwargs) -> Iterable[Iterable]:
        pass


    @abstractmethod
    def close(self) -> None:
        pass



class File(BaseFile, ABC):

    _File_Path: str = ""
    _Mode: str = ""
    _Encoding: str = ""

    @property
    def file_path(self) -> str:
        return self._File_Path


    @file_path.setter
    def file_path(self, path: str) -> None:
        self._File_Path = path


    @property
    def mode(self) -> str:
        return self._Mode


    @mode.setter
    def mode(self, mode: str) -> None:
        self._Mode = mode


    @property
    def encoding(self) -> str:
        return self._Encoding


    @encoding.setter
    def encoding(self, encoding: str) -> None:
        self._Encoding = encoding



class CSVFormat(File):

    _File_IO_Wrapper = None

    def open(self) -> None:
        self._File_IO_Wrapper = open(file=self.file_path, mode=self.mode, newline='', encoding=self.encoding)


    def write(self, data: Iterable[Iterable]) -> None:
        # Check format of data
        csv_data = CSVFormat._data_handling(data=data)
        # Write data
        csv_obj = csv.writer(self._File_IO_Wrapper)
        for data_line in csv_data:
            csv_obj.writerow(data_line)


    def read(self, delimiter: str = ",", dialect: str = None) -> Iterable[Iterable]:
        csv_data = csv.reader(self._File_IO_Wrapper, delimiter=delimiter, dialect=dialect)
        _data = [_data_row for _data_row in csv_data]
        return _data


    def close(self) -> None:
        self._File_IO_Wrapper.close()


    @staticmethod
    def _data_handling(data: List[list]) -> List[list]:
        checking_data = Checking.data(data=data)
        csv_data: List[list] = [d for d in checking_data]
        return csv_data



class XLSXFormat(File):

    __WorkBook = None
    __Sheet_Page = None

    def __init__(self, sheet_page: str = "sheet_page_1"):
        self.__Sheet_Page_Name = sheet_page


    def open(self) -> None:
        self.__WorkBook: Workbook = Workbook()
        self.__Sheet_Page = self.__WorkBook.create_sheet(index=0, title=self.__Sheet_Page_Name)


    def write(self, data: Iterable[Iterable]) -> None:
        for d in data:
            _handled_d = self._handle_data_row(data=d)
            self.__Sheet_Page.append(_handled_d)


    def _handle_data_row(self, data: Iterable) -> list:
        _handled_d = map(lambda a: str(a), data)
        return list(_handled_d)


    def read(self) -> Iterable[Iterable]:
        _workbook = load_workbook(filename=self.file_path, read_only=True)
        _sheet_page_data = []
        for _sheet in _workbook:
            for _row in _sheet.iter_rows(min_row=_sheet.min_row, max_row=_sheet.max_row, min_col=_sheet.min_column, max_col=_sheet.max_column):
                _data_row = [_cell.value for _cell in _row]
                _sheet_page_data.append(_data_row)
        return _sheet_page_data


    def close(self) -> None:
        self.__WorkBook.save(self.file_path)


    @staticmethod
    def _data_handling(data: List[list]) -> List[list]:
        checking_data = Checking.data(data=data)
        xlsx_data: List[list] = [d for d in checking_data]
        return xlsx_data



class JSONFormat(File):

    __JSON_IO = None

    def open(self) -> None:
        self.__JSON_IO = open(file=self.file_path, mode=self.mode, encoding=self.encoding)


    def write(self, data: Iterable[Iterable]) -> None:
        json_data = JSONFormat._data_handling(data=data)
        self.__JSON_IO.write(json_data)


    def read(self) -> Iterable[Iterable]:
        _data = json.load(self.__JSON_IO)
        return _data


    def close(self) -> None:
        self.__JSON_IO.close()


    @staticmethod
    def _data_handling(data: List[list]) -> str:
        json_data = json.dumps(data, ensure_ascii=False, default=str)
        return json_data



class XMLFormat(File):

    _File_IO_Wrapper = None

    def open(self) -> None:
        self._File_IO_Wrapper = open(file=self.file_path, mode=self.mode)


    def write(self, data: Iterable[Iterable]) -> None:
        _tasks_ele = ET.Element("tasks")
        for _data_row in data:
            _task_ele = ET.SubElement(_tasks_ele, "task")

            _url = ET.SubElement(_task_ele, "url")
            _url.text = str(_data_row[0])

            _http_method = ET.SubElement(_task_ele, "http-method")
            _http_method.text = str(_data_row[1])

        _tasks_xml_data = ET.tostring(element=_tasks_ele, encoding=self.encoding)
        self._File_IO_Wrapper.write(_tasks_xml_data)


    def read(self, *args, **kwargs) -> Iterable[Iterable]:
        _lines = self._File_IO_Wrapper.readline()
        _xml_root = ET.fromstringlist(_lines)

        _data = []
        for child in _xml_root:
            _data_row = self._parse_task(task=child)
            _data.append(_data_row)

        return _data


    def _parse_task(self, task: Element) -> list:
        return [_task_content.text for _task_content in task]


    def close(self) -> None:
        self._File_IO_Wrapper.close()



class PropertiesFormat(File):

    _File_IO_Wrapper = None
    __Config_Parser: configparser.ConfigParser = None

    def open(self) -> None:
        self._File_IO_Wrapper = open(file=self.file_path, mode=self.mode)
        self.__Config_Parser = configparser.ConfigParser()


    def write(self, data: Iterable[Iterable], sections: Iterable[str] = None) -> None:
        if sections is None:
            sections = [f"task_{i}" for i in range(1, len(data) + 1)]

        for _section, _data_row in zip(sections, data):
            self.__Config_Parser[_section] = {
                f"{_section}.url": str(_data_row[0]),
                f"{_section}.http_method": str(_data_row[1])
            }

        self.__Config_Parser.write(fp=self._File_IO_Wrapper)


    def read(self, *args, **kwargs) -> Iterable[Iterable]:
        _data = []
        self.__Config_Parser.read(filenames=self.file_path)
        _sections = self.__Config_Parser.sections()
        for _section in _sections:
            _config_content = list(self.__Config_Parser[_section].values())
            _data.append(_config_content)

        return _data


    def close(self) -> None:
        self._File_IO_Wrapper.close()



class Checking:

    def __init__(self):
        raise RuntimeError("All methods in class 'CheckingUtils' is static method, you shouldn't new this class.")


    @classmethod
    def data(cls, data: List[list]) -> Union[List[list], str]:
        """
        Description:
            Check the data format of data row is valid or invalid.
            It will raise DataRowFormatIsInvalidError if the data
            format is invalid.
        :param data:
        :return:
        """

        __checksum = map(cls.__is_data_row, data)
        if False in list(__checksum):
            # raise DataRowFormatIsInvalidError
            raise ValueError("")
        else:
            return list(data)


    @classmethod
    def __is_data_row(cls, data_row: Iterable) -> bool:
        """
        Description:
            First step: Checking first level of data format tree.
        :param data_row:
        :return:
        """

        if type(data_row) is list or type(data_row) is tuple:
            return cls.__is_data_content(data_row=data_row)
        else:
            return False


    @classmethod
    def __is_data_content(cls, data_row: Iterable) -> bool:
        """
        Description:
            First step: Checking second level of data format tree.
        :param data_row:
        :return:
        """

        chk_data_content = map(
            lambda row: False if isinstance(row, List) or isinstance(row, Tuple) else True,
            data_row)
        if False in list(chk_data_content):
            return False
        else:
            return True

